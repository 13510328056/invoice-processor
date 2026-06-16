"""
发票批处理工具 — Excel 输出模块

4 个 Sheet + 可选商品明细 Sheet5
原子写入策略：内存构建 → .tmp 写入 → os.rename 覆盖
"""

from __future__ import annotations

import logging
import os
import time
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

from src.config.loader import AppConfig
from src.models.enums import ProcessingStatus
from src.models.invoice import InvoiceData, LineItem
from src.models.result import ProcessingResult, ProcessingStatistics

logger = logging.getLogger(__name__)


# 成功 Sheet 的列头（22 列，与 InvoiceData.to_excel_row() 对应）
SUCCESS_HEADERS = [
    "文件路径", "发票代码", "发票号码", "开票日期", "校验码",
    "发票类型", "购买方名称", "购买方纳税人识别号", "购买方地址电话",
    "购买方开户行及账号", "销售方名称", "销售方纳税人识别号",
    "销售方地址电话", "销售方开户行及账号",
    "价税合计（大写）", "价税合计（小写）", "不含税金额", "税额",
    "商品明细 JSON", "处理时间", "MarkItDown 原始输出", "备注",
]

# 金额列索引（0-based）
AMOUNT_COLUMNS = [15, 16, 17]  # P=价税合计, Q=不含税金额, R=税额


class ExcelWriter:
    """Excel 输出写入器"""

    def __init__(self, config: AppConfig):
        self.config = config
        self.output_path = Path(config.output.filename)

    def write(
        self,
        results: list[ProcessingResult],
        stats: ProcessingStatistics,
    ) -> Path:
        """
        写入 Excel 文件（原子写入）

        Args:
            results: 处理结果列表
            stats: 处理统计

        Returns:
            输出文件的路径

        Raises:
            ExcelWriteError: 写入失败（重试后）
        """
        wb = Workbook()

        # ── Sheet1: 成功处理 ──
        ws_success = wb.active
        ws_success.title = self.config.output.sheet_success
        self._write_success_sheet(ws_success, results)

        # ── Sheet2: 失败处理 ──
        ws_failed = wb.create_sheet(self.config.output.sheet_failed)
        self._write_failed_sheet(ws_failed, results)

        # ── Sheet3: 非电子发票 ──
        ws_skipped = wb.create_sheet(self.config.output.sheet_skipped)
        self._write_skipped_sheet(ws_skipped, results)

        # ── Sheet4: 处理统计 ──
        ws_stats = wb.create_sheet(self.config.output.sheet_statistics)
        self._write_statistics_sheet(ws_stats, stats)

        # ── Sheet5: 商品明细（可选） ──
        self._write_line_items_sheet(wb, results)

        # ── 原子写入 ──
        return self._atomic_save(wb)

    def _write_success_sheet(self, ws, results: list[ProcessingResult]) -> None:
        """写入成功处理 Sheet"""
        # 表头
        ws.append(SUCCESS_HEADERS)
        ws.freeze_panes = "A2"

        # 数据行
        for result in results:
            if result.status == ProcessingStatus.SUCCESS and result.invoice:
                ws.append(result.invoice.to_excel_row())

        # U 列隐藏（MarkItDown 原始输出）
        ws.column_dimensions["U"].hidden = True

        # 金额列格式
        for col_idx in AMOUNT_COLUMNS:
            col_letter = get_column_letter(col_idx + 1)  # 1-based
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row,
                min_col=col_idx + 1, max_col=col_idx + 1,
            ):
                for cell in row:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        self._auto_column_width(ws)

    def _write_failed_sheet(self, ws, results: list[ProcessingResult]) -> None:
        """写入失败处理 Sheet"""
        ws.append(["文件路径", "失败原因", "处理时间"])
        ws.freeze_panes = "A2"
        for r in results:
            if r.status == ProcessingStatus.FAILED:
                ws.append([
                    r.scanned_file.rel_path,
                    r.error_message,
                    time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(r.scanned_file.modification_time)),
                ])
        self._auto_column_width(ws)

    def _write_skipped_sheet(self, ws, results: list[ProcessingResult]) -> None:
        """写入非电子发票 Sheet"""
        ws.append(["文件路径", "文件类型", "处理时间"])
        ws.freeze_panes = "A2"
        for r in results:
            if r.status == ProcessingStatus.SKIPPED and not r.invoice:
                ws.append([
                    r.scanned_file.rel_path,
                    r.scanned_file.file_type,
                    time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(r.scanned_file.modification_time)),
                ])
        self._auto_column_width(ws)

    def _write_statistics_sheet(self, ws, stats: ProcessingStatistics) -> None:
        """写入统计 Sheet（树状结构）"""
        ws.append(["统计项", "数量"])
        ws.append(["扫描文件总数", stats.total_scanned])
        ws.append(["├─ 电子发票文件数", stats.total_invoice_files])
        ws.append(["│  ├─ 识别成功", stats.total_success])
        ws.append(["│  └─ 识别失败", stats.total_failed])
        ws.append(["└─ 非电子发票文件数", stats.total_skipped])
        ws.append(["校验关系", self._validation_text(stats)])
        ws.append(["成功率", f"{stats.success_rate}%"])
        ws.append(["处理总耗时", self._format_duration(stats.total_wall_time)])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _write_line_items_sheet(self, wb: Workbook,
                                results: list[ProcessingResult]) -> None:
        """可选写入商品明细 Excel5"""
        # 从 config 判断是否展开（预留：可通过 config.yaml 控制）
        # 当前仅为 OFD 路径生成商品明细

        all_items: list[tuple[str, LineItem]] = []
        for r in results:
            if r.status == ProcessingStatus.SUCCESS and r.invoice:
                for item in r.invoice.line_items:
                    all_items.append((r.invoice.invoice_number, item))

        if not all_items:
            return

        ws = wb.create_sheet("商品明细")
        ws.append(["发票号码", "货物名称", "规格型号", "单位", "数量",
                    "单价", "金额", "税率", "税额"])
        ws.freeze_panes = "A2"

        for inv_no, item in all_items:
            ws.append([
                inv_no,
                item.goods_name,
                item.spec_model,
                item.unit,
                item.quantity,
                item.unit_price,
                item.amount,
                item.tax_rate,
                item.tax_amount,
            ])
        self._auto_column_width(ws)

    def _atomic_save(self, wb: Workbook) -> Path:
        """
        原子写入：写 .tmp → os.rename 覆盖

        Raises:
            ExcelWriteError: 重试后仍失败
        """
        target = self.output_path.resolve()
        tmp_path = target.with_name(f"~{target.name}.tmp")

        max_retries = self.config.processing.retry_count

        for attempt in range(max_retries + 1):
            try:
                # 写入临时文件
                wb.save(tmp_path)
                # 原子覆盖
                if target.exists():
                    target.unlink()  # Windows 上 rename 需要先删除
                tmp_path.rename(target)
                logger.info(f"Excel 输出成功: {target}")
                return target

            except PermissionError:
                if attempt < max_retries:
                    wait = (attempt + 1) * 2
                    logger.warning(
                        f"Excel 文件被占用，{wait}秒后重试 ({attempt + 1}/{max_retries})..."
                    )
                    time.sleep(wait)
                else:
                    logger.error(
                        f"Excel 写入失败: 文件被占用，请关闭 Excel 后重试。"
                        f"临时文件保留: {tmp_path}"
                    )
                    raise
            except Exception as e:
                # 清理临时文件
                if tmp_path.exists():
                    tmp_path.unlink()
                logger.error(f"Excel 写入异常: {e}")
                raise

        # 不应到达此处
        raise RuntimeError(f"Excel 写入失败: {target}")

    def _auto_column_width(self, ws) -> None:
        """自动调整列宽"""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    # 中文字符按 2 倍宽度计算
                    length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_length = max(max_length, length)
            adjusted = min(max_length + 2, 60)
            ws.column_dimensions[col_letter].width = adjusted

    @staticmethod
    def _validation_text(stats: ProcessingStatistics) -> str:
        """生成校验关系文本"""
        parts = []
        parts.append(f"{stats.total_scanned} = {stats.total_invoice_files} + {stats.total_skipped}")
        parts.append(f"{stats.total_invoice_files} = {stats.total_success} + {stats.total_failed}")
        return "；".join(parts) + (" ✓" if stats.validation_ok else " ✗")

    @staticmethod
    def _format_duration(seconds: float) -> str:
        """格式化时间"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
