"""
集成测试：端到端管线测试
"""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

import pytest

from src.config.loader import AppConfig, load_config
from src.models.enums import FileType, ProcessingStatus
from src.models.invoice import InvoiceData, LineItem
from src.models.result import ScannedFile, ProcessingResult, ProcessingStatistics
from src.output.excel_writer import ExcelWriter
from src.output.stat_reporter import aggregate
from src.pipeline.dedup import deduplicate
from src.pipeline.scanner import scan_directory, classify_files


class TestScannerIntegration:
    """扫描集成测试"""

    def test_mixed_directory(self, temp_with_files: str, sample_config: AppConfig):
        files = scan_directory(
            temp_with_files,
            skip_directories=sample_config.processing.skip_directories,
        )
        ofd, ocr, unsupported = classify_files(files)

        assert len(ofd) >= 1
        assert len(ocr) >= 3
        assert len(unsupported) >= 1


class TestDedupIntegration:
    """去重集成测试"""

    def test_dedup_on(self):
        """启用去重"""
        results = [
            ProcessingResult(
                scanned_file=ScannedFile("a.pdf", "a.pdf", "pdf"),
                status=ProcessingStatus.SUCCESS,
                invoice=InvoiceData(invoice_code="C1", invoice_number="N1"),
            ),
            ProcessingResult(
                scanned_file=ScannedFile("b.pdf", "b.pdf", "pdf"),
                status=ProcessingStatus.SUCCESS,
                invoice=InvoiceData(invoice_code="C1", invoice_number="N1"),
            ),
        ]
        deduped = deduplicate(results)
        success_count = sum(1 for r in deduped if r.status == ProcessingStatus.SUCCESS)
        assert success_count == 1


class TestStatisticsIntegration:
    """统计聚合集成测试"""

    def test_mixed_stats(self):
        results = [
            ProcessingResult(
                scanned_file=ScannedFile("/a.pdf", "a.pdf", "pdf"),
                status=ProcessingStatus.SUCCESS,
            ),
            ProcessingResult(
                scanned_file=ScannedFile("/b.pdf", "b.pdf", "jpg"),
                status=ProcessingStatus.FAILED,
            ),
            ProcessingResult(
                scanned_file=ScannedFile("/c.txt", "c.txt", "unsupported"),
                status=ProcessingStatus.SKIPPED,
            ),
        ]
        stats = aggregate(results)
        assert stats.total_scanned == 3
        assert stats.total_invoice_files == 2
        assert stats.total_success == 1
        assert stats.total_failed == 1
        assert stats.total_skipped == 1
        assert stats.validation_ok is True
        assert stats.success_rate == 50.0


class TestExcelWriterIntegration:
    """Excel 输出集成测试"""

    def test_write_success_sheet(self, tmp_path: Path, sample_config: AppConfig):
        sample_config.output.filename = str(tmp_path / "test_output.xlsx")

        results = [
            ProcessingResult(
                scanned_file=ScannedFile(
                    "/inv1.pdf", "inv1.pdf", "pdf",
                ),
                status=ProcessingStatus.SUCCESS,
                invoice=InvoiceData(
                    file_path="inv1.pdf",
                    invoice_code="C001",
                    invoice_number="N001",
                    total_amount="100.00",
                ),
                elapsed_seconds=1.5,
            ),
            ProcessingResult(
                scanned_file=ScannedFile(
                    "/inv2.ofd", "inv2.ofd", "ofd",
                ),
                status=ProcessingStatus.FAILED,
                error_message="解析失败",
                elapsed_seconds=2.0,
            ),
            ProcessingResult(
                scanned_file=ScannedFile(
                    "/readme.txt", "readme.txt", "unsupported",
                ),
                status=ProcessingStatus.SKIPPED,
            ),
        ]
        stats = aggregate(results)

        writer = ExcelWriter(sample_config)
        output = writer.write(results, stats)

        assert output.exists()
        assert output.suffix == ".xlsx"

        # 验证内容
        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "成功处理" in wb.sheetnames
        assert "失败处理" in wb.sheetnames
        assert "非电子发票" in wb.sheetnames
        assert "处理统计" in wb.sheetnames

        # 成功 Sheet 应有数据
        ws = wb["成功处理"]
        assert ws.max_row >= 2  # 表头 + 1 行数据
